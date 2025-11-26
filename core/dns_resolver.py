# core/dns_resolver.py
import subprocess
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from .utils import logger


def run_dns_resolution_and_export(merged_file: Path, result_dir: Path, input_identifier: str, dns_config: dict):
    command_template = dns_config.get("command", "").strip()
    if not command_template:
        raise ValueError("dns_resolution.command 不能为空")

    full_cmd = f"{command_template} -l {merged_file.absolute()}"
    logger.info(f"🚀 正在运行 DNS 解析: {full_cmd}")

    try:
        result = subprocess.run(
            full_cmd,
            shell=True,
            capture_output=True,
            text=True,
            timeout=300
        )
    except subprocess.TimeoutExpired:
        raise RuntimeError("dnsx 执行超时（5分钟）")

    if result.returncode != 0:
        logger.error(f"dnsx stderr: {result.stderr}")
        raise RuntimeError(f"dnsx 退出码 {result.returncode}")

    # === 初始化记录存储 ===
    records = {"A": [], "AAAA": [], "CNAME": [], "MX": [], "TXT": []}
    a_domains = set()

    # 编译 ANSI 清理正则（可复用）
    ansi_escape = re.compile(r'\x1B[@-_][0-?]*[ -/]*[@-~]')

    # === 逐行解析 dnsx 输出 ===
    for line in result.stdout.splitlines():
        line = line.strip()
        if not line:
            continue

        # 清理 ANSI 颜色码
        clean_line = ansi_escape.sub('', line)

        # 精准匹配: domain [TYPE] [value]
        match = re.match(r'^([^\s]+)\s+\[([A-Z]+)\]\s+\[(.*)\]$', clean_line)
        if not match:
            continue

        domain, rtype, value = match.groups()
        domain = domain.lower().rstrip('.')

        if rtype == "A":
            records["A"].append((domain, value))
            a_domains.add(domain)
        elif rtype == "AAAA":
            records["AAAA"].append((domain, value))
            a_domains.add(domain)
        elif rtype == "CNAME":
            records["CNAME"].append((domain, value))
        elif rtype == "MX":
            parts = value.split(maxsplit=1)
            priority = parts[0] if len(parts) > 0 else ""
            mail_server = parts[1] if len(parts) > 1 else value
            records["MX"].append((domain, priority, mail_server))
        elif rtype == "TXT":
            records["TXT"].append((domain, value))

    # === 写入 Excel ===
    timestamp = merged_file.stem.split('_')[-2:]
    timestamp_str = '_'.join(timestamp)
    excel_filename = f"{input_identifier}_dns_{timestamp_str}.xlsx"
    excel_path = result_dir / excel_filename

    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "Raw Merged"
    ws_raw.append(["Subdomain"])
    ws_raw.column_dimensions['A'].width = 40
    with open(merged_file, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            domain = line.strip().lower().rstrip('.')
            if domain:
                ws_raw.append([domain])

    for rtype in ["A", "AAAA", "CNAME", "MX", "TXT"]:
        if not records[rtype]:
            continue
        ws = wb.create_sheet(title=rtype)
        if rtype in ("A", "AAAA"):
            ws.append(["Subdomain", "IP"])
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 20
            for domain, ip in records[rtype]:
                ws.append([domain, ip])
        elif rtype == "CNAME":
            ws.append(["Subdomain", "Target"])
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 40
            for domain, target in records[rtype]:
                ws.append([domain, target])
        elif rtype == "MX":
            ws.append(["Subdomain", "Priority", "Mail Server"])
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 40
            for domain, priority, server in records[rtype]:
                ws.append([domain, priority, server])
        elif rtype == "TXT":
            ws.append(["Subdomain", "TXT Value"])
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 60
            for domain, txt in records[rtype]:
                ws.append([domain, txt])

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    wb.save(excel_path)
    logger.debug(f"✅ Excel 报告已保存: {excel_path}")

    # === 写入 reachable.txt ===
    reachable_filename = f"{input_identifier}_reachable.txt"
    reachable_path = result_dir / reachable_filename
    with open(reachable_path, 'w', encoding='utf-8') as f:
        for domain in sorted(a_domains):
            f.write(domain + '\n')
    logger.debug(f"✅ 可探测目标清单已保存: {reachable_path}")

    return excel_path, reachable_path